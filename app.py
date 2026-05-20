import re
import io
import os
import json
import subprocess
import tempfile
from flask import Flask, request, send_file, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# ── Column positions (pdftotext -layout output) ──
COMP_S, COMP_E = 19, 70
CONC_S, CONC_E = 70, 118
DEB_S,  DEB_E  = 118, 138
CRED_S, CRED_E = 138, 160
SALDO_S        = 160

DATE_RE   = re.compile(r'^ {0,5}(\d{2}/\d{2}/\d{4}) {2,}')
HEADER_RE = re.compile(r'COMPROBANTE.*CONCEPTO')

GARBAGE = [
    ' registrado en el Banco si,',
    ' Se presume conformidad',
    ' dentro de los sesenta',
    ' Banco Bica SA',
    'Inscripto - CUIT: 30-71233123',
    '01.04.008',
]

def safe(s, a, b=None):
    if b: return s[a:b].strip() if len(s) > a else ''
    return s[a:].strip() if len(s) > a else ''

def parse_num(s):
    s = s.strip()
    if not s: return None
    neg = s.startswith('(') and s.endswith(')')
    try: return (-1 if neg else 1) * float(s.strip('()').replace(',', ''))
    except: return None

def clean_concept(t):
    for g in GARBAGE:
        i = t.find(g)
        if i != -1: t = t[:i]
    return t.strip()

def parse_pdf(file_bytes):
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        result = subprocess.run(
            ['pdftotext', '-layout', tmp_path, '-'],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode != 0:
            raise RuntimeError('pdftotext falló: ' + result.stderr[:200])
        lines = result.stdout.split('\n')
    finally:
        os.unlink(tmp_path)

    rows, current, in_table = [], None, False

    for line in lines:
        if HEADER_RE.search(line):
            in_table = True
            continue
        if not in_table:
            continue
        stripped = line.strip()
        if not stripped or stripped == 'TRANSPORTE':
            continue

        m = DATE_RE.match(line)
        if m:
            if current:
                rows.append(current)
            current = {
                'fecha':    m.group(1),
                'comp':     safe(line, COMP_S, COMP_E),
                'concepto': safe(line, CONC_S, CONC_E),
                'debito':   parse_num(safe(line, DEB_S, DEB_E)),
                'credito':  parse_num(safe(line, CRED_S, SALDO_S)),
                'saldo':    parse_num(safe(line, SALDO_S)),
            }
        elif current and len(line) > CONC_S:
            extra = safe(line, CONC_S, CONC_E)
            if extra:
                current['concepto'] = (current['concepto'] + ' ' + extra).strip()

    if current:
        rows.append(current)

    for r in rows:
        r['concepto'] = clean_concept(r['concepto'])

    return rows

def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
    NORM_FILL   = PatternFill("solid", fgColor="FFFFFF")
    DEB_FILL    = PatternFill("solid", fgColor="FFE0E0")
    CRED_FILL   = PatternFill("solid", fgColor="E0F0E0")
    SALDO_FILL  = PatternFill("solid", fgColor="FFF8DC")
    CTAS_FILL   = PatternFill("solid", fgColor="FFFACD")
    thin        = Side(style='thin', color='CCCCCC')
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["Fecha", "Comprobante", "Concepto", "Débito", "Crédito", "Saldo", "Cuenta Contable"]
    col_widths = [12, 16, 80, 15, 15, 16, 25]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for i, row in enumerate(rows, start=2):
        fill = ALT_FILL if i % 2 == 0 else NORM_FILL
        data = [row['fecha'], row['comp'], row['concepto'],
                row['debito'], row['credito'], row['saldo'], '']

        for col, val in enumerate(data, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif col == 2:
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif col == 3:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            elif col in (4, 5, 6):
                cell.alignment = Alignment(horizontal='right', vertical='top')
                if val is not None:
                    cell.number_format = '#,##0.00'
                if col == 4 and val is not None:
                    cell.fill = DEB_FILL
                elif col == 5 and val is not None:
                    cell.fill = CRED_FILL
                elif col == 6:
                    cell.fill = SALDO_FILL
                    cell.font = Font(name="Arial", size=9, bold=True)
            elif col == 7:
                cell.fill = CTAS_FILL
                cell.alignment = Alignment(horizontal='left', vertical='top')

    ws.auto_filter.ref = f"A1:G{len(rows)+1}"

    ws2 = wb.create_sheet("Resumen")
    ws2['A1'] = "Resumen del Extracto"
    ws2['A1'].font = Font(bold=True, size=14, name="Arial", color="1F4E79")

    total_deb  = sum(r['debito']  for r in rows if r['debito']  is not None)
    total_cred = sum(r['credito'] for r in rows if r['credito'] is not None)
    saldo_ini  = rows[0]['saldo']  if rows else 0
    saldo_fin  = rows[-1]['saldo'] if rows else 0

    summary = [
        ("Total movimientos",   len(rows)),
        ("Movimientos débito",  sum(1 for r in rows if r['debito']  is not None)),
        ("Movimientos crédito", sum(1 for r in rows if r['credito'] is not None)),
        ("", ""),
        ("Saldo inicial",  saldo_ini),
        ("Total débitos",  total_deb),
        ("Total créditos", total_cred),
        ("Saldo final",    saldo_fin),
    ]
    for r_idx, (label, val) in enumerate(summary, start=3):
        c1 = ws2.cell(row=r_idx, column=1, value=label)
        c2 = ws2.cell(row=r_idx, column=2, value=val)
        c1.font = Font(name="Arial", size=10, bold=bool(label))
        c2.font = Font(name="Arial", size=10)
        if isinstance(val, float):
            c2.number_format = '#,##0.00'
            c2.alignment = Alignment(horizontal='right')
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BICA PDF → Excel</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Sora:wght@300;400;600&display=swap');
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  :root {
    --navy: #0f1c2e; --navy2: #1a2f48; --accent: #2a9d8f;
    --light: #f0f4f8; --text: #1a2f48; --muted: #6b7e96;
    --border: #d0dbe8; --white: #ffffff;
    --green-text: #0f7a6b; --red-text: #c0392b;
    --red: #fde8e8;
  }
  body { font-family: 'Sora', sans-serif; background: var(--light); color: var(--text); min-height: 100vh; display: flex; flex-direction: column; align-items: center; padding: 2rem 1rem 4rem; }

  #splash { position: fixed; inset: 0; background: var(--navy); display: flex; flex-direction: column; align-items: center; justify-content: center; gap: 2rem; z-index: 999; transition: opacity 0.5s ease; }
  #splash.hidden { opacity: 0; pointer-events: none; }
  .splash-icon { width: 72px; height: 72px; background: rgba(42,157,143,0.15); border: 2px solid var(--accent); border-radius: 20px; display: flex; align-items: center; justify-content: center; }
  .splash-icon svg { width: 36px; height: 36px; fill: none; stroke: var(--accent); stroke-width: 2; stroke-linecap: round; stroke-linejoin: round; }
  .splash-title { font-size: 1.6rem; font-weight: 600; color: var(--white); letter-spacing: -0.02em; text-align: center; }
  .splash-sub { font-size: 0.85rem; color: rgba(255,255,255,0.45); font-family: 'DM Mono', monospace; text-align: center; margin-top: 6px; }
  #startBtn { padding: 0.9rem 3rem; background: var(--accent); color: var(--white); border: none; border-radius: 12px; font-family: 'Sora', sans-serif; font-size: 1rem; font-weight: 600; cursor: pointer; min-width: 200px; transition: background 0.15s, transform 0.1s; }
  #startBtn:hover { background: #238f82; }
  #startBtn:disabled { background: #1a5c55; cursor: default; }
  .splash-status { font-size: 0.78rem; font-family: 'DM Mono', monospace; color: rgba(255,255,255,0.35); min-height: 1.2em; text-align: center; }
  .splash-status.ok  { color: var(--accent); }
  .splash-status.err { color: #e76f51; }
  .dot-anim::after { content: ''; animation: dots 1.2s steps(4,end) infinite; }
  @keyframes dots { 0%{content:''} 25%{content:'.'} 50%{content:'..'} 75%{content:'...'} }

  #app { width: 100%; display: flex; flex-direction: column; align-items: center; }
  header { width: 100%; max-width: 680px; margin-bottom: 2.5rem; }
  .logo { display: flex; align-items: center; gap: 12px; margin-bottom: 0.4rem; }
  .logo-icon { width: 40px; height: 40px; background: var(--navy); border-radius: 10px; display: flex; align-items: center; justify-content: center; }
  .logo-icon svg { width: 22px; height: 22px; fill: var(--accent); }
  h1 { font-size: 1.5rem; font-weight: 600; color: var(--navy); letter-spacing: -0.02em; }
  .subtitle { font-size: 0.85rem; color: var(--muted); font-weight: 300; margin-top: 2px; }
  .card { width: 100%; max-width: 680px; background: var(--white); border: 1px solid var(--border); border-radius: 16px; padding: 2rem; }
  .drop-zone { border: 2px dashed var(--border); border-radius: 12px; padding: 3rem 2rem; text-align: center; cursor: pointer; transition: all 0.2s; position: relative; background: var(--light); }
  .drop-zone:hover, .drop-zone.drag-over { border-color: var(--accent); background: #eaf7f5; }
  .drop-zone input[type="file"] { position: absolute; inset: 0; opacity: 0; cursor: pointer; width: 100%; height: 100%; }
  .drop-icon { width: 52px; height: 52px; margin: 0 auto 1rem; background: var(--navy); border-radius: 12px; display: flex; align-items: center; justify-content: center; }
  .drop-icon svg { width: 28px; height: 28px; fill: none; stroke: var(--accent); stroke-width: 2; stroke-linecap: round; stroke-linejoin: round; }
  .drop-label { font-size: 1rem; font-weight: 600; color: var(--navy); margin-bottom: 0.3rem; }
  .drop-hint { font-size: 0.8rem; color: var(--muted); font-family: 'DM Mono', monospace; }
  .file-info { display: none; margin-top: 1.2rem; padding: 0.75rem 1rem; background: var(--light); border-radius: 8px; border: 1px solid var(--border); align-items: center; gap: 10px; font-size: 0.85rem; }
  .file-info.visible { display: flex; }
  .file-name { font-family: 'DM Mono', monospace; font-weight: 500; color: var(--navy); flex: 1; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .file-size { color: var(--muted); font-size: 0.78rem; }
  .btn { display: block; width: 100%; margin-top: 1.5rem; padding: 0.9rem; background: var(--navy); color: var(--white); border: none; border-radius: 10px; font-family: 'Sora', sans-serif; font-size: 0.95rem; font-weight: 600; cursor: pointer; transition: background 0.15s, transform 0.1s; }
  .btn:hover { background: var(--navy2); }
  .btn:disabled { background: var(--border); color: var(--muted); cursor: not-allowed; }
  .progress-wrap { display: none; margin-top: 1.5rem; }
  .progress-wrap.visible { display: block; }
  .progress-label { font-size: 0.82rem; color: var(--muted); margin-bottom: 0.5rem; font-family: 'DM Mono', monospace; display: flex; justify-content: space-between; }
  .progress-bar-bg { height: 6px; background: var(--light); border-radius: 99px; overflow: hidden; border: 1px solid var(--border); }
  .progress-bar-fill { height: 100%; background: linear-gradient(90deg, var(--accent), #48cfc0); border-radius: 99px; width: 0%; transition: width 0.3s ease; }
  .result { display: none; margin-top: 1.5rem; border-radius: 10px; overflow: hidden; border: 1px solid var(--border); }
  .result.visible { display: block; }
  .result-header { background: var(--navy); padding: 0.9rem 1.2rem; display: flex; align-items: center; gap: 10px; }
  .result-header svg { width: 18px; height: 18px; fill: var(--accent); flex-shrink: 0; }
  .result-title { font-size: 0.9rem; font-weight: 600; color: var(--white); flex: 1; }
  .stats { display: grid; grid-template-columns: repeat(3,1fr); }
  .stat { padding: 1rem 1.2rem; border-right: 1px solid var(--border); }
  .stat:last-child { border-right: none; }
  .stat-label { font-size: 0.72rem; color: var(--muted); font-family: 'DM Mono', monospace; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px; }
  .stat-value { font-size: 1.1rem; font-weight: 600; font-family: 'DM Mono', monospace; color: var(--navy); }
  .stat-value.green { color: var(--green-text); }
  .stat-value.red   { color: var(--red-text); }
  .download-btn { display: flex; align-items: center; justify-content: center; gap: 8px; width: calc(100% - 2.4rem); margin: 1rem 1.2rem 1.2rem; padding: 0.75rem; background: var(--accent); color: var(--white); border: none; border-radius: 8px; font-family: 'Sora', sans-serif; font-size: 0.9rem; font-weight: 600; cursor: pointer; text-decoration: none; transition: background 0.15s; }
  .download-btn:hover { background: #238f82; }
  .download-btn svg { width: 16px; height: 16px; fill: none; stroke: currentColor; stroke-width: 2; stroke-linecap: round; stroke-linejoin: round; }
  .error { display: none; margin-top: 1.2rem; padding: 0.85rem 1rem; background: var(--red); border: 1px solid #f5c6c6; border-radius: 8px; font-size: 0.85rem; color: var(--red-text); font-family: 'DM Mono', monospace; }
  .error.visible { display: block; }
  .note { max-width: 680px; margin-top: 1.5rem; font-size: 0.78rem; color: var(--muted); text-align: center; line-height: 1.6; font-family: 'DM Mono', monospace; }
</style>
</head>
<body>

<div id="splash">
  <div class="splash-icon">
    <svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/></svg>
  </div>
  <div>
    <div class="splash-title">BICA PDF → Excel</div>
    <div class="splash-sub">Extracto bancario · Banco BICA SA</div>
  </div>
  <button id="startBtn" onclick="iniciar()">Iniciar</button>
  <div class="splash-status" id="splashStatus"></div>
</div>

<div id="app" style="display:none">
  <header>
    <div class="logo">
      <div class="logo-icon">
        <svg viewBox="0 0 24 24"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/></svg>
      </div>
      <div>
        <h1>BICA PDF → Excel</h1>
        <div class="subtitle">Extracto bancario · Banco BICA SA</div>
      </div>
    </div>
  </header>

  <div class="card">
    <div class="drop-zone" id="dropZone">
      <input type="file" id="fileInput" accept=".pdf" />
      <div class="drop-icon">
        <svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>
      </div>
      <div class="drop-label">Arrastrá el PDF acá</div>
      <div class="drop-hint">o hacé click para seleccionar · solo .pdf</div>
    </div>
    <div class="file-info" id="fileInfo">
      <svg style="width:16px;height:16px;flex-shrink:0" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/></svg>
      <span class="file-name" id="fileName"></span>
      <span class="file-size" id="fileSize"></span>
    </div>
    <div class="error" id="errorBox"></div>
    <button class="btn" id="convertBtn" disabled>Convertir a Excel</button>
    <div class="progress-wrap" id="progressWrap">
      <div class="progress-label">
        <span id="progressText">Procesando PDF...</span>
        <span id="progressPct">0%</span>
      </div>
      <div class="progress-bar-bg"><div class="progress-bar-fill" id="progressFill"></div></div>
    </div>
    <div class="result" id="result">
      <div class="result-header">
        <svg viewBox="0 0 24 24"><path d="M9 11l3 3L22 4"/><path d="M21 12v7a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2V5a2 2 0 0 1 2-2h11"/></svg>
        <span class="result-title">Excel generado correctamente</span>
      </div>
      <div class="stats">
        <div class="stat"><div class="stat-label">Movimientos</div><div class="stat-value" id="statTotal">—</div></div>
        <div class="stat"><div class="stat-label">Total créditos</div><div class="stat-value green" id="statCred">—</div></div>
        <div class="stat"><div class="stat-label">Total débitos</div><div class="stat-value red" id="statDeb">—</div></div>
      </div>
      <a class="download-btn" id="downloadBtn" href="#" download>
        <svg viewBox="0 0 24 24"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="7 10 12 15 17 10"/><line x1="12" y1="15" x2="12" y2="3"/></svg>
        Descargar Excel
      </a>
    </div>
  </div>
  <div class="note">Compatible con extractos del Banco BICA SA (formato Reporting Services)</div>
</div>

<script>
async function iniciar() {
  const btn = document.getElementById('startBtn');
  const status = document.getElementById('splashStatus');
  btn.disabled = true;
  status.className = 'splash-status dot-anim';
  status.textContent = 'Iniciando servidor';
  for (let i = 0; i < 20; i++) {
    try {
      const r = await fetch('/ping', { cache: 'no-store' });
      if (r.ok) {
        status.className = 'splash-status ok';
        status.textContent = '✓ Servidor listo';
        await new Promise(res => setTimeout(res, 600));
        const splash = document.getElementById('splash');
        splash.classList.add('hidden');
        document.getElementById('app').style.display = 'flex';
        setTimeout(() => splash.remove(), 600);
        return;
      }
    } catch(e) {}
    await new Promise(res => setTimeout(res, 2000));
  }
  status.className = 'splash-status err';
  status.textContent = 'No se pudo conectar. Reintentá.';
  btn.disabled = false;
  btn.textContent = 'Reintentar';
}

function fmtSize(b) { return b < 1048576 ? (b/1024).toFixed(0)+' KB' : (b/1048576).toFixed(1)+' MB'; }
function fmtNum(n) { return new Intl.NumberFormat('es-AR',{minimumFractionDigits:2}).format(n); }

const dropZone   = document.getElementById('dropZone');
const fileInput  = document.getElementById('fileInput');
const convertBtn = document.getElementById('convertBtn');
const progressWrap = document.getElementById('progressWrap');
const progressFill = document.getElementById('progressFill');
const progressText = document.getElementById('progressText');
const progressPct  = document.getElementById('progressPct');
const resultEl   = document.getElementById('result');
const errorBox   = document.getElementById('errorBox');
const downloadBtn= document.getElementById('downloadBtn');
let selectedFile = null;

function setFile(file) {
  if (!file || !file.name.toLowerCase().endsWith('.pdf')) { showError('Solo se aceptan archivos .pdf'); return; }
  selectedFile = file;
  document.getElementById('fileName').textContent = file.name;
  document.getElementById('fileSize').textContent = fmtSize(file.size);
  document.getElementById('fileInfo').classList.add('visible');
  convertBtn.disabled = false;
  resultEl.classList.remove('visible');
  errorBox.classList.remove('visible');
}
function showError(msg) { errorBox.textContent = msg; errorBox.classList.add('visible'); }
function animateProgress(target, duration, label) {
  progressText.textContent = label;
  const start = parseFloat(progressFill.style.width) || 0;
  const t0 = performance.now();
  function step(now) {
    const t = Math.min((now-t0)/duration, 1);
    const e = t<0.5?2*t*t:-1+(4-2*t)*t;
    const cur = start + (target-start)*e;
    progressFill.style.width = cur+'%';
    progressPct.textContent = Math.round(cur)+'%';
    if (t<1) requestAnimationFrame(step);
  }
  requestAnimationFrame(step);
}

fileInput.addEventListener('change', e => { if (e.target.files[0]) setFile(e.target.files[0]); });
dropZone.addEventListener('dragover', e => { e.preventDefault(); dropZone.classList.add('drag-over'); });
dropZone.addEventListener('dragleave', () => dropZone.classList.remove('drag-over'));
dropZone.addEventListener('drop', e => { e.preventDefault(); dropZone.classList.remove('drag-over'); if (e.dataTransfer.files[0]) setFile(e.dataTransfer.files[0]); });

convertBtn.addEventListener('click', async () => {
  if (!selectedFile) return;
  errorBox.classList.remove('visible');
  resultEl.classList.remove('visible');
  convertBtn.disabled = true;
  progressWrap.classList.add('visible');
  progressFill.style.width = '0%';
  animateProgress(30, 400, 'Leyendo PDF...');
  const fd = new FormData();
  fd.append('file', selectedFile);
  setTimeout(() => animateProgress(65, 1500, 'Extrayendo movimientos...'), 500);
  try {
    const resp = await fetch('/convert', { method: 'POST', body: fd });
    if (!resp.ok) {
      const err = await resp.json().catch(() => ({}));
      throw new Error(err.error || 'Error en el servidor (HTTP '+resp.status+')');
    }
    animateProgress(90, 400, 'Generando Excel...');
    const stats = JSON.parse(resp.headers.get('X-Stats') || '{}');
    const blob  = await resp.blob();
    setTimeout(() => {
      animateProgress(100, 300, 'Listo!');
      setTimeout(() => {
        progressWrap.classList.remove('visible');
        convertBtn.disabled = false;
        const url = URL.createObjectURL(blob);
        downloadBtn.href = url;
        downloadBtn.download = selectedFile.name.replace(/\.pdf$/i,'')+'_Movimientos.xlsx';
        document.getElementById('statTotal').textContent = stats.total || '—';
        document.getElementById('statCred').textContent  = stats.creditos ? fmtNum(stats.creditos) : '—';
        document.getElementById('statDeb').textContent   = stats.debitos  ? fmtNum(stats.debitos)  : '—';
        resultEl.classList.add('visible');
        downloadBtn.click();
      }, 400);
    }, 600);
  } catch(err) {
    progressWrap.classList.remove('visible');
    convertBtn.disabled = false;
    showError('Error: ' + err.message);
  }
});
</script>
</body>
</html>
"""

@app.route('/ping')
def ping():
    return jsonify({'status': 'ok'})

@app.route('/')
def index():
    return HTML

@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return jsonify({'error': 'No se recibió archivo'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'El archivo debe ser un PDF'}), 400
    try:
        rows = parse_pdf(f.read())
        if not rows:
            return jsonify({'error': 'No se encontraron movimientos. ¿Es un extracto del Banco BICA?'}), 400
        excel_buf  = build_excel(rows)
        total_deb  = sum(r['debito']  for r in rows if r['debito']  is not None)
        total_cred = sum(r['credito'] for r in rows if r['credito'] is not None)
        stats = json.dumps({'total': len(rows), 'debitos': total_deb, 'creditos': total_cred})
        response = send_file(excel_buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name='movimientos.xlsx')
        response.headers['X-Stats'] = stats
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5050))
    app.run(host='0.0.0.0', port=port, debug=False)
